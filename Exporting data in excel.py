from typing import Union, Any
from ultralytics import YOLO
import cv2
import cvzone
import math
from sort import *
import os
import openpyxl



cap = cv2.VideoCapture("../Videos/3 way traffic.mp4")  # For Video
model = YOLO("../Yolo-Weights/yolov8l.pt")


classNames = ["person", "bicycle", "car", "motorbike", "aeroplane", "bus", "train", "truck", "boat",
              "traffic light", "fire hydrant", "stop sign", "parking meter", "bench", "bird", "cat",
              "dog", "horse", "sheep", "cow", "elephant", "bear", "zebra", "giraffe", "backpack", "umbrella",
              "handbag", "tie", "suitcase", "frisbee", "skis", "snowboard", "sports ball", "kite", "baseball bat",
              "baseball glove", "skateboard", "surfboard", "tennis racket", "bottle", "wine glass", "cup",
              "fork", "knife", "spoon", "bowl", "banana", "apple", "sandwich", "orange", "broccoli",
              "carrot", "hot dog", "pizza", "donut", "cake", "chair", "sofa", "pottedplant", "bed",
              "diningtable", "toilet", "tvmonitor", "laptop", "mouse", "remote", "keyboard", "cell phone",
              "microwave", "oven", "toaster", "sink", "refrigerator", "book", "clock", "vase", "scissors",
              "teddy bear", "hair drier", "toothbrush"
              ]





mask = cv2.imread("mask.png")
tracker = Sort(max_age=20, min_hits=3, iou_threshold=0.3)  # creating instance




l1 = [245, 300, 500, 300]
l2 = [550, 270, 655, 270]
l3 = [805, 500, 1180, 500]

totalCount_l1 = []
totalCount_l2 = []
totalCount_l3 = []

confidence_l1 = []
confidence_l2 = []
confidence_l3 =[]

class_l1 = []
class_l2 = []
class_l3 = []





while True:
    success, img = cap.read()
    if not success:
        break
    imgRegion = cv2.bitwise_and(img, mask)
    results = model(imgRegion, stream=True)
    detections = np.empty((0, 5))



    for r in results:
        boxes = r.boxes
        for box in boxes:
            x1, y1, x2, y2 = box.xyxy[0]
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
            w, h = x2 - x1, y2 - y1


            conf: float = math.ceil((box.conf[0] * 100)) / 100
            cls = int(box.cls[0])
            currentClass: Union[str, Any] = classNames[cls]






            if currentClass == "car" or currentClass == "truck" or currentClass == "bus" or currentClass == "motorbike" and conf > 0.3:
                currentArray = np.array([x1, y1, x2, y2, conf])
                detections = np.vstack((detections, currentArray))
                cvzone.putTextRect(img, f' {currentClass}', (x1 + 10, y1 + 20), colorT=(255, 255, 255), colorR=(60, 20, 220), font=cv2.FONT_HERSHEY_TRIPLEX, scale=0.75, thickness=1, offset=2)
    resultsTracker = tracker.update(detections)




    cv2.line(img, (l1[0], l1[1]), (l1[2], l1[3]), (0, 0, 255), 5)
    cv2.line(img, (l2[0], l2[1]), (l2[2], l2[3]), (0, 0, 255), 5)
    cv2.line(img, (l3[0], l3[1]), (l3[2], l3[3]), (0, 0, 255), 5)






    for result in resultsTracker:
        x1, y1, x2, y2, id = result
        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
        # print(result)
        w, h = x2 - x1, y2 - y1
        cvzone.cornerRect(img, (x1, y1, w, h), l=17, t=2, rt=1, colorR=(255, 228, 181), colorC=(60, 20, 220))






        cx, cy = x1 + w // 2, y1 + h // 2
        cv2.circle(img, (cx, cy), 5, (255, 0, 255), cv2.FILLED)




        if l1[0] < cx < l1[2] and l1[1] - 15 < cy < l1[1] + 15:
            if totalCount_l1.count(id) == 0:
                totalCount_l1.append(id)
                confidence_l1.append(conf)
                class_l1.append(currentClass)
                cv2.line(img, (l1[0], l1[1]), (l1[2], l1[3]), (0, 255, 0), 5)
        elif l2[0] < cx < l2[2] and l2[1] - 15 < cy < l2[1] + 15:
            if totalCount_l2.count(id) == 0:
                totalCount_l2.append(id)
                confidence_l2.append(conf)
                class_l2.append(currentClass)
                cv2.line(img, (l2[0], l2[1]), (l2[2], l2[3]), (0, 255, 0), 5)
        elif l3[0] < cx < l3[2] and l3[1]-15 < cy < l3[1]+15:
            if totalCount_l3.count(id) == 0:
                totalCount_l3.append(id)
                confidence_l3.append(conf)
                class_l3.append(currentClass)
                cv2.line(img, (l3[0], l3[1]), (l3[2], l3[3]), (0, 255, 0), 5)



    cv2.putText(img, f'Total Count L1: {len(totalCount_l1)}', (10, 40), cv2.FONT_HERSHEY_PLAIN, 2, (0, 0, 0), 2)
    cv2.putText(img, f'Total Count L2: {len(totalCount_l2)}', (10, 80), cv2.FONT_HERSHEY_PLAIN, 2, (0, 0, 0), 2)
    cv2.putText(img, f'Total Count L3: {len(totalCount_l3)}', (10, 120), cv2.FONT_HERSHEY_PLAIN, 2, (0, 0, 0), 2)





    cv2.imshow("Image", img)
    # cv2.imshow("ImageRegion", imgRegion)
    cv2.waitKey(1)



# Specify the file name
filename = 'vehicular_data.xlsx'
# Check if the file exists
if not os.path.exists(filename):
    # If the file doesn't exist, create a new workbook and save it
    book = openpyxl.Workbook()
    book.save(filename)
else:
    # If the file exists, load the existing workbook
    book = openpyxl.load_workbook(filename)


# Now we can work with 'book' as your workbook


# Create two sheets
# Check if the sheets exist, if not create them
sheet1 = book.get_sheet_by_name("way_1") if "way_1" in book.sheetnames else book.create_sheet("way_1")
sheet2 = book.get_sheet_by_name("way_2") if "way_2" in book.sheetnames else book.create_sheet("way_2")
sheet3 = book.get_sheet_by_name("way_3") if "way_3" in book.sheetnames else book.create_sheet("way_3")


# Write the list names to the first row of each column in Sheet1
sheet1.cell(row=2, column=2, value='Vehicle_id')
sheet1.cell(row=2, column=3, value='Class_name')
sheet1.cell(row=2, column=4, value='Confidence')


# Write each list to a different column in Sheet1, starting from the third row
for i in range(len(totalCount_l1)):
    sheet1.cell(row=i+3, column=2, value=totalCount_l1[i])
    sheet1.cell(row=i+3, column=3, value=class_l1[i])
    sheet1.cell(row=i+3, column=4, value=confidence_l1[i])

# Write the list names to the first row of each column in Sheet2
sheet2.cell(row=2, column=2, value='Vehicle_id')
sheet2.cell(row=2, column=3, value='Class_name')
sheet2.cell(row=2, column=4, value='Confidence')


# Write each list to a different column in Sheet2, starting from the third row
for i in range(len(totalCount_l2)):
    sheet2.cell(row=i+3, column=2, value=totalCount_l2[i])
    sheet2.cell(row=i+3, column=3, value=class_l2[i])
    sheet2.cell(row=i+3, column=4, value=confidence_l2[i])


# Write the list names to the first row of each column in Sheet3
sheet3.cell(row=2, column=2, value='Vehicle_id')
sheet3.cell(row=2, column=3, value='Class_name')
sheet3.cell(row=2, column=4, value='Confidence')


# Write each list to a different column in Sheet3, starting from the third row
for i in range(len(totalCount_l3)):
    sheet3.cell(row=i+3, column=2, value=totalCount_l3[i])
    sheet3.cell(row=i+3, column=3, value=class_l3[i])
    sheet3.cell(row=i+3, column=4, value=confidence_l3[i])

book.save(filename)

print("Operation Successful \n Check the excel file in the current working directory")
